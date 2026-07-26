"""Spreadsheet export for assembled study topics."""

from __future__ import annotations

import re
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from app.storage import OutputStorage

_SOURCE_LINK = re.compile(
    r"^- \[(?P<title>.+?)]\((?P<url>https?://[^)]+)\)$",
    re.MULTILINE,
)

# Excel's hard per-cell character limit. openpyxl does not enforce or warn
# about this -- it silently truncates on save -- so a thoroughly researched
# "deep" section (which easily runs past this many characters once its
# source list is included) would otherwise lose content with no indication.
_EXCEL_CELL_CHAR_LIMIT = 32767
_TRUNCATION_NOTICE = "\n\n[...이하 생략: 엑셀 셀 글자 수 제한. 전체 내용은 Markdown 다운로드 참고]"


def _fit_cell_text(text: str) -> str:
    if len(text) <= _EXCEL_CELL_CHAR_LIMIT:
        return text
    truncated_length = _EXCEL_CELL_CHAR_LIMIT - len(_TRUNCATION_NOTICE)
    return text[:truncated_length] + _TRUNCATION_NOTICE


def _excel_safe(value: Any) -> str:
    """Remove characters that cannot be serialized in an XML workbook."""
    return ILLEGAL_CHARACTERS_RE.sub("", str(value))


def _prepare_sheet(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_excel_workbook(topic: str, storage: OutputStorage) -> BytesIO:
    """Build an xlsx workbook from the persisted TOC and section documents."""

    toc: list[dict[str, Any]] = storage.read_json(storage.toc_json_path)
    manifest = storage.load_manifest()
    state_by_id = {
        str(section.get("id")): section for section in manifest.get("sections", [])
    }

    workbook = Workbook()
    toc_sheet = workbook.active
    toc_sheet.title = "목차"
    body_sheet = workbook.create_sheet("본문")
    source_sheet = workbook.create_sheet("출처")

    _prepare_sheet(toc_sheet, ["섹션 ID", "제목", "설명"])
    _prepare_sheet(body_sheet, ["섹션 ID", "제목", "본문"])
    _prepare_sheet(source_sheet, ["섹션 ID", "출처 제목", "URL"])

    for section in toc:
        section_id = _excel_safe(section["id"])
        title = _excel_safe(section["title"])
        toc_sheet.append(
            [section_id, title, _excel_safe(section.get("description", ""))]
        )

        state = state_by_id.get(section_id, {})
        section_path = storage.topic_dir / str(state.get("path", ""))
        content = (
            section_path.read_text(encoding="utf-8")
            if state.get("status") == "done" and section_path.is_file()
            else ""
        )
        safe_content = _excel_safe(content)
        body_sheet.append([section_id, title, _fit_cell_text(safe_content)])
        body_sheet.cell(row=body_sheet.max_row, column=3).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        for match in _SOURCE_LINK.finditer(content):
            source_sheet.append(
                [
                    section_id,
                    _excel_safe(match.group("title")),
                    _excel_safe(match.group("url")),
                ]
            )

    toc_sheet.column_dimensions["A"].width = 12
    toc_sheet.column_dimensions["B"].width = 36
    toc_sheet.column_dimensions["C"].width = 72
    body_sheet.column_dimensions["A"].width = 12
    body_sheet.column_dimensions["B"].width = 36
    body_sheet.column_dimensions["C"].width = 100
    source_sheet.column_dimensions["A"].width = 12
    source_sheet.column_dimensions["B"].width = 48
    source_sheet.column_dimensions["C"].width = 72
    workbook.properties.title = _excel_safe(topic)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_section_zip(topic: str, storage: OutputStorage) -> BytesIO:
    """Build a ZIP containing the TOC and each completed section in TOC order."""
    del topic  # Kept in the public signature for parity with the other exporters.
    toc: list[dict[str, Any]] = storage.read_json(storage.toc_json_path)
    manifest = storage.load_manifest()
    state_by_id = {
        str(section.get("id")): section for section in manifest.get("sections", [])
    }

    output = BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "00-목차.md",
            storage.toc_markdown_path.read_text(encoding="utf-8"),
        )
        for section in toc:
            state = state_by_id.get(str(section["id"]), {})
            if state.get("status") != "done":
                continue
            relative_path = Path(str(state.get("path", "")))
            section_path = storage.topic_dir / relative_path
            archive.writestr(
                relative_path.name,
                section_path.read_text(encoding="utf-8"),
            )
    output.seek(0)
    return output
