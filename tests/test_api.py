import asyncio
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import app.jobs as jobs_module
from app.config import Settings
from app.jobs import SerialJobQueue
from app.main import create_app
from app.storage import OutputStorage
from app.toc import toc_to_markdown


class FakeQueue:
    def __init__(self, output_root: Path | None = None) -> None:
        self.started = False
        self.output_root = output_root
        self.section_jobs: list[tuple[str, str, bool]] = []
        self.build_jobs: list[tuple[str, list[str] | None, bool]] = []
        self.toc_jobs: list[tuple[str, str, int | None]] = []

    async def start(self) -> None:
        self.started = True

    async def stop(self) -> None:
        self.started = False

    async def enqueue_section(
        self,
        topic: str,
        section_id: str,
        *,
        force: bool = False,
    ) -> None:
        self.section_jobs.append((topic, section_id, force))

    async def enqueue_build(
        self,
        topic: str,
        *,
        sections_filter: list[str] | None = None,
        force_regenerate: bool = False,
    ) -> None:
        self.build_jobs.append((topic, sections_filter, force_regenerate))

    async def enqueue_toc_generation(
        self,
        topic: str,
        *,
        depth: str,
        num_sections: int | None,
    ) -> None:
        self.toc_jobs.append((topic, depth, num_sections))
        if self.output_root is not None:
            await fake_toc_generator_factory(self.output_root)(
                topic,
                depth=depth,
                num_sections=num_sections,
                output_root=self.output_root,
            )


def fake_toc_generator_factory(output_root: Path):
    async def fake_toc_generator(
        topic: str,
        *,
        depth: str,
        num_sections: int | None,
        output_root: Path,
    ) -> dict[str, Any]:
        count = num_sections or (10 if depth == "deep" else 6)
        sections = [
            {
                "id": f"{index:02d}",
                "title": f"Section {index}",
                "description": f"Scope {index}",
                "subsections": [
                    {
                        "id": f"{index:02d}.1",
                        "title": f"Part {index}",
                        "description": f"Part scope {index}",
                    }
                ],
            }
            for index in range(1, count + 1)
        ]
        storage = OutputStorage(output_root, topic)
        storage.write_json(storage.toc_json_path, sections)
        storage.write_text(
            storage.toc_markdown_path,
            toc_to_markdown(topic, sections),
        )
        manifest = storage.initialize_manifest(depth=depth, sections=sections)
        return {
            "toc": sections,
            "toc_path": str(storage.toc_json_path),
            "manifest": manifest,
        }

    return fake_toc_generator


@pytest.fixture
def api_client(tmp_path: Path) -> tuple[TestClient, FakeQueue, Settings]:
    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    queue = FakeQueue(tmp_path)
    application = create_app(
        settings=settings,
        job_queue=queue,
    )
    with TestClient(application) as client:
        yield client, queue, settings


def test_all_topic_routes(api_client: tuple[TestClient, FakeQueue, Settings]) -> None:
    client, queue, settings = api_client

    created = client.post(
        "/api/topics",
        json={"topic": "API Test Topic", "depth": "standard", "num_sections": 2},
    )
    assert created.status_code == 202
    slug = created.json()["slug"]
    created_manifest = OutputStorage(
        settings.research_output_dir, "API Test Topic"
    ).load_manifest()

    topics = client.get("/api/topics")
    assert topics.status_code == 200
    assert topics.json() == [
        {
            "topic": "API Test Topic",
            "slug": slug,
            "depth": "standard",
            "created_at": created_manifest["created_at"],
            "updated_at": created_manifest["updated_at"],
            "toc_status": "done",
            "completed_sections": 0,
            "total_sections": 2,
            "has_study_document": False,
        }
    ]

    detail = client.get(f"/api/topics/{slug}")
    assert detail.status_code == 200
    assert len(detail.json()["toc"]) == 2
    assert detail.json()["manifest"]["sections"][0]["status"] == "pending"

    section = client.post(f"/api/topics/{slug}/sections/01/research")
    assert section.status_code == 202
    assert section.json() == {"status": "queued"}
    assert queue.section_jobs == [("API Test Topic", "01", False)]

    build = client.post(
        f"/api/topics/{slug}/build",
        json={"sections_filter": ["02"], "force_regenerate": True},
    )
    assert build.status_code == 202
    assert build.json() == {"status": "queued"}
    assert queue.build_jobs == [("API Test Topic", ["02"], True)]

    storage = OutputStorage(settings.research_output_dir, "API Test Topic")
    storage.write_text(storage.study_document_path, "# Finished document")
    manifest = storage.load_manifest()
    manifest["study_document"] = {"path": "study_document.md"}
    storage.save_manifest(manifest)
    first_section = manifest["sections"][0]
    storage.write_text(
        storage.topic_dir / first_section["path"],
        "# Section 1\n\nBody text.\n\n## Sources\n\n"
        "- [Example source](https://example.com/source)",
    )
    storage.update_section("01", status="done", source_count=1)

    document = client.get(f"/api/topics/{slug}/document")
    assert document.status_code == 200
    assert document.text == "# Finished document\n"
    assert document.headers["content-type"] == "text/markdown; charset=utf-8"

    download = client.get(f"/api/topics/{slug}/download")
    assert download.status_code == 200
    assert download.content == b"# Finished document\n"
    assert download.headers["content-type"] == "text/markdown; charset=utf-8"
    assert "attachment" in download.headers["content-disposition"]

    markdown_download = client.get(f"/api/topics/{slug}/download?format=markdown")
    assert markdown_download.content == download.content
    assert markdown_download.headers["content-type"] == download.headers["content-type"]

    excel_download = client.get(f"/api/topics/{slug}/download?format=excel")
    assert excel_download.status_code == 200
    assert excel_download.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        f"filename*=utf-8''{slug}.xlsx"
        in excel_download.headers["content-disposition"]
    )
    workbook = load_workbook(BytesIO(excel_download.content))
    assert workbook.sheetnames == ["목차", "본문", "출처"]
    assert workbook["목차"]["B2"].value == "Section 1"
    assert "Body text." in workbook["본문"]["C2"].value
    assert workbook["본문"]["C2"].alignment.wrap_text is True
    assert workbook["출처"]["B2"].value == "Example source"
    assert workbook["출처"]["C2"].value == "https://example.com/source"

    zip_download = client.get(f"/api/topics/{slug}/download?format=zip")
    assert zip_download.status_code == 200
    assert zip_download.headers["content-type"] == "application/zip"
    assert (
        f"filename*=utf-8''{slug}.zip" in zip_download.headers["content-disposition"]
    )

    assert client.get(f"/api/topics/{slug}/download?format=pdf").status_code == 422

    deleted = client.delete(f"/api/topics/{slug}")
    assert deleted.status_code == 204
    assert not storage.topic_dir.exists()
    assert client.get(f"/api/topics/{slug}").status_code == 404


def test_excel_and_zip_downloads_encode_korean_filenames(tmp_path: Path) -> None:
    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    storage = OutputStorage(tmp_path, "한글 주제")
    sections = [{"id": "01", "title": "첫 장", "description": "설명"}]
    storage.write_json(storage.toc_json_path, sections)
    storage.write_text(storage.toc_markdown_path, "# 목차")
    manifest = storage.initialize_manifest(depth="standard", sections=sections)
    storage.write_text(
        storage.topic_dir / manifest["sections"][0]["path"],
        "# 첫 장\n\n본문",
    )
    storage.update_section("01", status="done")
    storage.write_text(storage.study_document_path, "# 완성 문서")

    with TestClient(create_app(settings=settings, job_queue=FakeQueue())) as client:
        excel = client.get(f"/api/topics/{storage.topic_slug}/download?format=excel")
        section_zip = client.get(
            f"/api/topics/{storage.topic_slug}/download?format=zip"
        )

    assert excel.status_code == 200
    assert (
        excel.headers["content-disposition"]
        == "attachment; filename*=utf-8''%ED%95%9C%EA%B8%80-%EC%A3%BC%EC%A0%9C.xlsx"
    )
    assert load_workbook(BytesIO(excel.content))["본문"]["B2"].value == "첫 장"
    assert section_zip.status_code == 200
    assert (
        section_zip.headers["content-disposition"]
        == "attachment; filename*=utf-8''%ED%95%9C%EA%B8%80-%EC%A3%BC%EC%A0%9C.zip"
    )


def test_api_reports_not_found_and_conflict(
    api_client: tuple[TestClient, FakeQueue, Settings],
) -> None:
    client, _, settings = api_client
    assert client.get("/api/topics/missing").status_code == 404

    created = client.post(
        "/api/topics",
        json={"topic": "Conflict Topic", "depth": "standard", "num_sections": 2},
    )
    slug = created.json()["slug"]
    assert client.post(
        "/api/topics",
        json={"topic": "Conflict Topic", "depth": "standard", "num_sections": 2},
    ).status_code == 409

    storage = OutputStorage(settings.research_output_dir, "Conflict Topic")
    storage.update_section("01", status="in_progress")
    duplicate = client.post(f"/api/topics/{slug}/sections/01/research")
    assert duplicate.status_code == 409
    assert client.post(f"/api/topics/{slug}/build", json={}).status_code == 409
    assert client.delete(f"/api/topics/{slug}").status_code == 409
    assert (
        client.post(f"/api/topics/{slug}/sections/99/research").status_code == 404
    )


def test_create_topic_returns_pending_state_without_waiting_for_toc(
    tmp_path: Path,
) -> None:
    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    queue = FakeQueue()
    with TestClient(create_app(settings=settings, job_queue=queue)) as client:
        created = client.post(
            "/api/topics",
            json={"topic": "Queued Topic", "depth": "deep", "num_sections": 8},
        )
        slug = created.json()["slug"]
        detail = client.get(f"/api/topics/{slug}")
        blocked_delete = client.delete(f"/api/topics/{slug}")

    assert created.status_code == 202
    assert created.json() == {"slug": "queued-topic", "status": "queued"}
    assert queue.toc_jobs == [("Queued Topic", "deep", 8)]
    assert detail.status_code == 200
    assert detail.json()["toc"] == []
    assert detail.json()["manifest"]["toc_status"] == "generating"
    assert blocked_delete.status_code == 409


def test_app_startup_reconciles_stale_toc_and_section_jobs(tmp_path: Path) -> None:
    generating = OutputStorage(tmp_path, "Interrupted TOC")
    generating.initialize_pending_manifest(depth="standard")
    researching = OutputStorage(tmp_path, "Interrupted Research")
    researching.initialize_manifest(
        depth="standard",
        sections=[{"id": "01", "title": "Section"}],
    )
    researching.write_json(
        researching.toc_json_path,
        [{"id": "01", "title": "Section"}],
    )
    researching.update_section("01", status="in_progress")

    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    with TestClient(create_app(settings=settings, job_queue=FakeQueue())) as client:
        toc_detail = client.get(f"/api/topics/{generating.topic_slug}").json()
        section_detail = client.get(f"/api/topics/{researching.topic_slug}").json()
        assert client.delete(f"/api/topics/{generating.topic_slug}").status_code == 204

    assert toc_detail["manifest"]["toc_status"] == "error"
    assert toc_detail["manifest"]["toc_error"] == "서버 재시작으로 목차 생성이 중단됨"
    assert section_detail["manifest"]["sections"][0]["status"] == "pending"


def test_research_section_rejects_redo_of_done_section_without_force(
    api_client: tuple[TestClient, FakeQueue, Settings],
) -> None:
    client, queue, settings = api_client
    created = client.post(
        "/api/topics",
        json={"topic": "Done Section Topic", "depth": "standard", "num_sections": 2},
    )
    slug = created.json()["slug"]
    storage = OutputStorage(settings.research_output_dir, "Done Section Topic")
    storage.update_section("01", status="done")

    # Without force: rejected, and no job is queued -- clicking an already
    # "done" section must not silently re-spend API budget on a full redo.
    rejected = client.post(f"/api/topics/{slug}/sections/01/research")
    assert rejected.status_code == 409
    assert queue.section_jobs == []

    # With force=true: allowed, and the queue receives the force flag.
    forced = client.post(f"/api/topics/{slug}/sections/01/research?force=true")
    assert forced.status_code == 202
    assert queue.section_jobs == [("Done Section Topic", "01", True)]


def test_get_section_document_only_returns_completed_sections(
    api_client: tuple[TestClient, FakeQueue, Settings],
) -> None:
    client, _, settings = api_client
    created = client.post(
        "/api/topics",
        json={"topic": "Section Document Topic", "depth": "standard", "num_sections": 2},
    )
    slug = created.json()["slug"]
    storage = OutputStorage(settings.research_output_dir, "Section Document Topic")
    manifest = storage.update_section("01", status="done")
    completed_section = manifest["sections"][0]
    storage.write_text(
        storage.section_path("01", completed_section["title"]),
        "# 완료된 섹션\n\n한국어 본문",
    )

    completed = client.get(f"/api/topics/{slug}/sections/01")
    assert completed.status_code == 200
    assert completed.text == "# 완료된 섹션\n\n한국어 본문\n"
    assert completed.headers["content-type"] == "text/markdown; charset=utf-8"

    assert client.get(f"/api/topics/{slug}/sections/02").status_code == 404
    assert client.get(f"/api/topics/{slug}/sections/99").status_code == 404


def test_get_section_document_uses_manifest_path_after_title_drift(
    api_client: tuple[TestClient, FakeQueue, Settings],
) -> None:
    client, _, settings = api_client
    created = client.post(
        "/api/topics",
        json={"topic": "Drifted API Topic", "depth": "standard", "num_sections": 2},
    )
    slug = created.json()["slug"]
    storage = OutputStorage(settings.research_output_dir, "Drifted API Topic")
    manifest = storage.load_manifest()
    section = manifest["sections"][0]
    actual_path = storage.topic_dir / section["path"]
    storage.write_text(actual_path, "# Body under the original filename")
    section["title"] = "Renamed Metadata Title"
    section["status"] = "done"
    storage.save_manifest(manifest)

    response = client.get(f"/api/topics/{slug}/sections/01")

    assert response.status_code == 200
    assert response.text == "# Body under the original filename\n"


def test_basic_auth_protects_api_and_static_frontend(tmp_path: Path) -> None:
    settings = Settings(
        require_api_key=False,
        research_output_dir=tmp_path,
        site_password="correct horse battery staple",
    )
    application = create_app(settings=settings, job_queue=FakeQueue())
    with TestClient(application) as client:
        assert client.get("/").status_code == 401
        assert client.get("/api/topics").status_code == 401
        assert client.get("/api/topics", auth=("researcher", "wrong")).status_code == 401

        authorized = client.get(
            "/api/topics",
            auth=("researcher", "correct horse battery staple"),
        )
        assert authorized.status_code == 200
        assert authorized.json() == []
        assert client.get(
            "/",
            auth=("researcher", "correct horse battery staple"),
        ).status_code == 200


def test_static_frontend_is_served_without_password(
    api_client: tuple[TestClient, FakeQueue, Settings],
) -> None:
    client, _, _ = api_client
    assert client.get("/").status_code == 200
    assert client.get("/app.js").status_code == 200
    assert client.get("/style.css").status_code == 200


@pytest.mark.asyncio
async def test_job_queue_never_runs_research_in_parallel(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storage = OutputStorage(tmp_path, "Serial Topic")
    sections = [
        {"id": "01", "title": "One"},
        {"id": "02", "title": "Two"},
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.initialize_manifest(depth="standard", sections=sections)
    active = 0
    maximum_active = 0

    async def fake_research(
        topic: str,
        section_id: str,
        **_kwargs: Any,
    ) -> dict[str, Any]:
        nonlocal active, maximum_active
        active += 1
        maximum_active = max(maximum_active, active)
        await asyncio.sleep(0.01)
        storage.update_section(section_id, status="done")
        active -= 1
        return {}

    monkeypatch.setattr(jobs_module, "research_section", fake_research)
    queue = SerialJobQueue(
        Settings(require_api_key=False, research_output_dir=tmp_path)
    )
    await queue.start()
    try:
        await queue.enqueue_section("Serial Topic", "01")
        await queue.enqueue_section("Serial Topic", "02")
        await queue.join()
    finally:
        await queue.stop()

    assert maximum_active == 1
    assert [
        section["status"] for section in storage.load_manifest()["sections"]
    ] == ["done", "done"]


@pytest.mark.asyncio
async def test_job_queue_timeout_marks_error_and_processes_next_job(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storage = OutputStorage(tmp_path, "Timeout Topic")
    sections = [
        {"id": "01", "title": "Hangs"},
        {"id": "02", "title": "Continues"},
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.initialize_manifest(depth="standard", sections=sections)

    async def fake_research(
        topic: str,
        section_id: str,
        **_kwargs: Any,
    ) -> dict[str, Any]:
        if section_id == "01":
            await asyncio.sleep(1)
        storage.update_section(section_id, status="done")
        return {}

    monkeypatch.setattr(jobs_module, "research_section", fake_research)
    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    settings.section_timeout_seconds = 0.01
    queue = SerialJobQueue(settings)
    await queue.start()
    try:
        await queue.enqueue_section("Timeout Topic", "01")
        await queue.enqueue_section("Timeout Topic", "02")
        await queue.join()
    finally:
        await queue.stop()

    assert [
        section["status"] for section in storage.load_manifest()["sections"]
    ] == ["error", "done"]


@pytest.mark.asyncio
async def test_job_queue_generates_toc_with_queued_options(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storage = OutputStorage(tmp_path, "Queued TOC")
    pending = storage.initialize_pending_manifest(depth="deep")
    calls: list[tuple[str, str, int | None, Path]] = []

    async def fake_generate_toc(
        topic: str,
        *,
        depth: str,
        num_sections: int | None,
        output_root: Path,
    ) -> dict[str, Any]:
        calls.append((topic, depth, num_sections, output_root))
        return storage.initialize_manifest(
            depth=depth,
            sections=[{"id": "01", "title": "Generated"}],
            created_at=pending["created_at"],
        )

    monkeypatch.setattr(jobs_module, "generate_toc", fake_generate_toc)
    queue = SerialJobQueue(
        Settings(require_api_key=False, research_output_dir=tmp_path)
    )
    await queue.start()
    try:
        await queue.enqueue_toc_generation(
            "Queued TOC", depth="deep", num_sections=7
        )
        await queue.join()
    finally:
        await queue.stop()

    assert calls == [("Queued TOC", "deep", 7, tmp_path)]
    assert storage.load_manifest()["toc_status"] == "done"


@pytest.mark.asyncio
@pytest.mark.parametrize("times_out", [False, True])
async def test_job_queue_marks_toc_error_on_failure_or_timeout(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    times_out: bool,
) -> None:
    topic = "TOC Timeout" if times_out else "TOC Failure"
    storage = OutputStorage(tmp_path, topic)
    storage.initialize_pending_manifest(depth="standard")

    async def fake_generate_toc(*_args: Any, **_kwargs: Any) -> dict[str, Any]:
        if times_out:
            await asyncio.sleep(1)
        raise RuntimeError("generation failed")

    monkeypatch.setattr(jobs_module, "generate_toc", fake_generate_toc)
    settings = Settings(require_api_key=False, research_output_dir=tmp_path)
    if times_out:
        settings.toc_timeout_seconds = 0.01
    queue = SerialJobQueue(settings)
    await queue.start()
    try:
        await queue.enqueue_toc_generation(
            topic, depth="standard", num_sections=None
        )
        await queue.join()
    finally:
        await queue.stop()

    manifest = storage.load_manifest()
    assert manifest["toc_status"] == "error"
    assert manifest["toc_error"] == ("" if times_out else "generation failed")


@pytest.mark.asyncio
async def test_build_skips_assembly_when_a_section_completes_with_no_sources(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A section that "succeeds" with zero sources must still block assembly.

    research_section() now marks such sections status="error" (DESIGN.md
    §22) instead of "done", so the existing "assemble only if every target
    is done" rule (§20.5) already covers this case for free -- this test
    just confirms the build-level wiring still holds for that path.
    """
    storage = OutputStorage(tmp_path, "No Sources Build")
    sections = [
        {"id": "01", "title": "Has Sources"},
        {"id": "02", "title": "Empty Context"},
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.initialize_manifest(depth="standard", sections=sections)
    assembled: list[str] = []

    async def fake_research(
        topic: str,
        section_id: str,
        **_kwargs: Any,
    ) -> dict[str, Any]:
        if section_id == "02":
            storage.update_section(section_id, status="error", source_count=0)
            return {"content_markdown": "No context was available.", "sources": []}
        storage.update_section(section_id, status="done", source_count=3)
        return {"content_markdown": "Real content.", "sources": [{"url": "https://example.com"}]}

    def fake_assemble(topic: str, **_kwargs: Any) -> dict[str, Any]:
        assembled.append(topic)
        return {}

    monkeypatch.setattr(jobs_module, "research_section", fake_research)
    monkeypatch.setattr(jobs_module, "assemble_study_document", fake_assemble)
    queue = SerialJobQueue(
        Settings(require_api_key=False, research_output_dir=tmp_path)
    )
    await queue.start()
    try:
        await queue.enqueue_build("No Sources Build")
        await queue.join()
    finally:
        await queue.stop()

    assert [
        section["status"] for section in storage.load_manifest()["sections"]
    ] == ["done", "error"]
    assert assembled == []


@pytest.mark.asyncio
async def test_build_researches_with_configured_concurrency_then_assembles(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    storage = OutputStorage(tmp_path, "Parallel Build")
    sections = [
        {"id": f"{index:02d}", "title": f"Section {index}"}
        for index in range(1, 5)
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.initialize_manifest(depth="standard", sections=sections)
    active = 0
    maximum_active = 0
    assembled: list[str] = []

    async def fake_research(
        topic: str,
        section_id: str,
        **_kwargs: Any,
    ) -> dict[str, Any]:
        nonlocal active, maximum_active
        active += 1
        maximum_active = max(maximum_active, active)
        try:
            await asyncio.sleep(0.02)
            storage.update_section(section_id, status="done")
            return {}
        finally:
            active -= 1

    def fake_assemble(topic: str, **_kwargs: Any) -> dict[str, Any]:
        assembled.append(topic)
        return {}

    monkeypatch.setattr(jobs_module, "research_section", fake_research)
    monkeypatch.setattr(jobs_module, "assemble_study_document", fake_assemble)
    queue = SerialJobQueue(
        Settings(
            require_api_key=False,
            research_output_dir=tmp_path,
            max_concurrent_research=2,
        )
    )
    await queue.start()
    try:
        await queue.enqueue_build("Parallel Build")
        await queue.join()
    finally:
        await queue.stop()

    assert maximum_active == 2
    assert [
        section["status"] for section in storage.load_manifest()["sections"]
    ] == ["done", "done", "done", "done"]
    assert assembled == ["Parallel Build"]


@pytest.mark.asyncio
async def test_build_continues_after_failure_and_skips_assembly(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
) -> None:
    storage = OutputStorage(tmp_path, "Partial Failure")
    sections = [
        {"id": f"{index:02d}", "title": f"Section {index}"}
        for index in range(1, 4)
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.initialize_manifest(depth="standard", sections=sections)
    assembled: list[str] = []

    async def fake_research(
        topic: str,
        section_id: str,
        **_kwargs: Any,
    ) -> dict[str, Any]:
        await asyncio.sleep(0.01)
        if section_id == "02":
            raise RuntimeError("deliberate section failure")
        storage.update_section(section_id, status="done")
        return {}

    def fake_assemble(topic: str, **_kwargs: Any) -> dict[str, Any]:
        assembled.append(topic)
        return {}

    monkeypatch.setattr(jobs_module, "research_section", fake_research)
    monkeypatch.setattr(jobs_module, "assemble_study_document", fake_assemble)
    queue = SerialJobQueue(
        Settings(
            require_api_key=False,
            research_output_dir=tmp_path,
            max_concurrent_research=2,
        )
    )
    await queue.start()
    try:
        await queue.enqueue_build("Partial Failure")
        await queue.join()
    finally:
        await queue.stop()

    assert [
        section["status"] for section in storage.load_manifest()["sections"]
    ] == ["done", "error", "done"]
    assert assembled == []
    assert "Skipping document assembly" in caplog.text
