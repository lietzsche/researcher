from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from app.export import (
    _EXCEL_CELL_CHAR_LIMIT,
    build_excel_workbook,
    build_section_zip,
)
from app.storage import OutputStorage


def _make_topic(tmp_path: Path, *, section_body: str) -> OutputStorage:
    storage = OutputStorage(tmp_path, "Excel Limits")
    sections = [{"id": "01", "title": "Section 1", "description": "desc"}]
    storage.write_json(storage.toc_json_path, sections)
    manifest = storage.initialize_manifest(depth="standard", sections=sections)
    section_path = storage.topic_dir / manifest["sections"][0]["path"]
    storage.write_text(section_path, section_body)
    storage.update_section("01", status="done", source_count=1)
    return storage


def test_build_excel_workbook_keeps_short_body_intact(tmp_path: Path) -> None:
    storage = _make_topic(tmp_path, section_body="short body")

    workbook = load_workbook(build_excel_workbook(storage.topic, storage))

    assert workbook["본문"]["C2"].value.strip() == "short body"


def test_build_excel_workbook_truncates_bodies_over_the_cell_limit(
    tmp_path: Path,
) -> None:
    long_body = "x" * (_EXCEL_CELL_CHAR_LIMIT + 5000) + "\n\n## Sources\n\n- [S](https://example.com)"
    storage = _make_topic(tmp_path, section_body=long_body)

    workbook = load_workbook(build_excel_workbook(storage.topic, storage))

    cell_value = workbook["본문"]["C2"].value
    assert len(cell_value) <= _EXCEL_CELL_CHAR_LIMIT
    assert "엑셀 셀 글자 수 제한" in cell_value
    # Sources are extracted from the full, untruncated content.
    assert workbook["출처"]["C2"].value == "https://example.com"


def test_build_excel_workbook_removes_illegal_xml_characters(tmp_path: Path) -> None:
    storage = _make_topic(
        tmp_path,
        section_body=(
            "body\x00with\x0billegal chars\n\n## Sources\n\n"
            "- [source\x0ctitle](https://example.com/\x01path)"
        ),
    )
    toc = storage.read_json(storage.toc_json_path)
    toc[0]["title"] = "title\x02text"
    toc[0]["description"] = "description\x03text"
    storage.write_json(storage.toc_json_path, toc)
    storage.topic = "topic\x04title"

    workbook = load_workbook(build_excel_workbook(storage.topic, storage))

    assert workbook.properties.title == "topictitle"
    assert workbook["목차"]["B2"].value == "titletext"
    assert workbook["목차"]["C2"].value == "descriptiontext"
    assert "bodywithillegal chars" in workbook["본문"]["C2"].value
    assert workbook["출처"]["B2"].value == "sourcetitle"
    assert workbook["출처"]["C2"].value == "https://example.com/path"


def test_build_section_zip_includes_only_done_sections_in_toc_order(
    tmp_path: Path,
) -> None:
    storage = OutputStorage(tmp_path, "ZIP Topic")
    sections = [
        {"id": "01", "title": "첫 번째"},
        {"id": "02", "title": "두 번째"},
        {"id": "03", "title": "세 번째"},
    ]
    storage.write_json(storage.toc_json_path, sections)
    storage.write_text(storage.toc_markdown_path, "# 목차")
    manifest = storage.initialize_manifest(depth="standard", sections=sections)
    for index in (0, 2):
        state = manifest["sections"][index]
        storage.write_text(storage.topic_dir / state["path"], f"# {state['title']}")
        storage.update_section(state["id"], status="done")

    with ZipFile(build_section_zip(storage.topic, storage)) as archive:
        assert archive.namelist() == [
            "00-목차.md",
            Path(manifest["sections"][0]["path"]).name,
            Path(manifest["sections"][2]["path"]).name,
        ]
        assert archive.read("00-목차.md").decode() == "# 목차\n"
        assert "두 번째" not in "\n".join(archive.namelist())
